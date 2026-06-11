"""Processamento de uploads: extração de texto, persistência em /tmp/uploads e TTL.

Cada upload gera dois arquivos:
  /tmp/uploads/{uuid}        -> bytes originais
  /tmp/uploads/{uuid}.json   -> metadados (nome, tipo, texto extraído ou base64)
"""

import base64
import io
import json
import os
import time
import uuid

UPLOAD_DIR = "/tmp/uploads"
TTL_MINUTES = int(os.environ.get("UPLOAD_TTL_MINUTES", "60"))

IMAGE_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}
DOC_EXTENSIONS = {".pdf", ".txt", ".md", ".docx", ".xlsx", ".csv", ".json"}

# Assinaturas de conteúdo (magic bytes) — barra arquivo malicioso disfarçado
# (ex.: executável renomeado para .png). Whitelist por extensão + conteúdo real.
_MAGIC = {
    ".png": [b"\x89PNG\r\n\x1a\n"],
    ".jpg": [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".gif": [b"GIF87a", b"GIF89a"],
    ".webp": [b"RIFF"],
    ".pdf": [b"%PDF"],
    ".docx": [b"PK\x03\x04"],  # zip
    ".xlsx": [b"PK\x03\x04"],  # zip
}


def _validate_content(ext: str, data: bytes) -> None:
    """Confere se o conteúdo bate com a extensão declarada. Levanta ValueError."""
    sigs = _MAGIC.get(ext)
    if sigs is not None:
        if not any(data.startswith(s) for s in sigs):
            raise ValueError(
                f"Conteúdo do arquivo não corresponde à extensão {ext} (possível arquivo disfarçado)"
            )
        if ext == ".webp" and data[8:12] != b"WEBP":
            raise ValueError("Conteúdo do arquivo não corresponde à extensão .webp")
    else:
        # Tipos texto (.txt .md .csv .json): rejeita binário disfarçado (bytes NUL)
        if b"\x00" in data[:8192]:
            raise ValueError(f"Arquivo {ext} contém dados binários (esperado texto puro)")


def _meta_path(file_id: str) -> str:
    return os.path.join(UPLOAD_DIR, f"{file_id}.json")


def _data_path(file_id: str) -> str:
    return os.path.join(UPLOAD_DIR, file_id)


def _extract_pdf(data: bytes) -> str:
    import pdfplumber

    parts = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                parts.append(text)
    return "\n\n".join(parts)


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(data: bytes, max_rows: int = 200) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(f"... (truncado em {max_rows} linhas)")
                break
            rows.append("\t".join("" if c is None else str(c) for c in row))
        parts.append(f"--- Planilha: {sheet.title} ---\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def process_upload(filename: str, data: bytes) -> dict:
    """Processa um arquivo enviado e persiste em disco. Retorna metadados públicos."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(filename)[1].lower()
    file_id = str(uuid.uuid4())

    meta = {
        "file_id": file_id,
        "name": os.path.basename(filename),
        "created_at": time.time(),
    }

    if ext in IMAGE_TYPES:
        _validate_content(ext, data)
        meta["kind"] = "image"
        meta["media_type"] = IMAGE_TYPES[ext]
        meta["base64"] = base64.standard_b64encode(data).decode("utf-8")
        meta["preview"] = f"data:{meta['media_type']};base64,{meta['base64']}"
    elif ext in DOC_EXTENSIONS:
        _validate_content(ext, data)
        meta["kind"] = "document"
        try:
            if ext == ".pdf":
                text = _extract_pdf(data)
            elif ext == ".docx":
                text = _extract_docx(data)
            elif ext == ".xlsx":
                text = _extract_xlsx(data)
            else:  # .txt / .md / .csv / .json
                text = data.decode("utf-8", errors="replace")
        except ValueError:
            raise
        except Exception:
            raise ValueError(f"Arquivo {ext} corrompido ou inválido (falha ao extrair conteúdo)")
        meta["text"] = text
        meta["preview"] = text[:300]
    else:
        raise ValueError(f"Tipo de arquivo não suportado: {ext or filename}")

    with open(_data_path(file_id), "wb") as f:
        f.write(data)
    with open(_meta_path(file_id), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)

    return {
        "file_id": file_id,
        "name": meta["name"],
        "type": meta["kind"],
        "preview": meta["preview"],
    }


def load_file_meta(file_id: str) -> dict | None:
    path = _meta_path(file_id)
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def cleanup_expired_uploads() -> None:
    if not os.path.isdir(UPLOAD_DIR):
        return
    cutoff = time.time() - TTL_MINUTES * 60
    for entry in os.listdir(UPLOAD_DIR):
        path = os.path.join(UPLOAD_DIR, entry)
        try:
            if os.path.isfile(path) and os.path.getmtime(path) < cutoff:
                os.remove(path)
        except OSError:
            pass
